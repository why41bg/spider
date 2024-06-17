from csv import writer
from os.path import getsize
from pathlib import Path
from platform import system
from sqlite3 import connect
from time import localtime
from time import strftime

from openpyxl import Workbook
from openpyxl import load_workbook

from src.customizer import WARNING, ERROR, INFO
from src.stringcleaner import Cleaner

__all__ = [
    'BaseLogger',
    'NoneLogger',
    'CSVLogger',
    'XLSXLogger',
    'SQLLogger',
    'RecordManager']


class BaseLogger:
    """不记录日志，空白日志记录器"""

    def __init__(self, main_path: Path, console, root="", folder="", name=""):
        self.log = None  # 记录器主体
        self.console = console
        self._root, self._folder, self._name = self.init_check(
            main_path=main_path,
            root=root,
            folder=folder,
            name=name,
        )

    def init_check(
            self,
            main_path: Path,
            root=None,
            folder=None,
            name=None) -> tuple:
        root = self.check_root(root, main_path)
        folder = self.check_folder(folder)
        name = self.check_name(name)
        return root, folder, name

    def check_root(self, root: str, default: Path) -> Path:
        if (r := Path(root)).is_dir():
            return r
        if root:
            self.console.print(f"日志储存路径 {root} 无效，程序将使用项目根路径作为储存路径")
        return default

    def check_name(self, name: str) -> str:
        if not name:
            return "%Y-%m-%d %H.%M.%S"
        try:
            _ = strftime(name, localtime())
            return name
        except ValueError:
            self.console.print(f"日志名称格式 {name} 无效，程序将使用默认时间格式：年-月-日 时.分.秒")
            return "%Y-%m-%d %H.%M.%S"

    @staticmethod
    def check_folder(folder: str) -> str:
        return Cleaner().filter_name(folder, False, "Log")

    def run(self, *args, **kwargs):
        pass

    def info(self, text: str, output=True, **kwargs):
        if output:
            self.console.print(text, style=INFO, **kwargs)

    def warning(self, text: str, output=True, **kwargs):
        if output:
            self.console.print(text, style=WARNING, **kwargs)

    def error(self, text: str, output=True, **kwargs):
        if output:
            self.console.print(text, style=ERROR, **kwargs)


class NoneLogger:
    def __init__(self, *args, **kwargs):
        self.field_keys = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def save(self, *args, **kwargs):
        pass

    @staticmethod
    def _rename(root: Path, type_: str, old: str, new_: str) -> str:
        mark = new_.split("_", 1)
        if not old or mark[-1] == old:
            return new_
        mark[-1] = old
        old_file = root.joinpath(f'{"_".join(mark)}.{type_}')
        if old_file.exists():
            new_file = root.joinpath(f"{new_}.{type_}")
            old_file.rename(new_file)
        return new_


class CSVLogger(NoneLogger):
    """CSV格式记录"""
    __type = "csv"

    def __init__(
            self,
            root: Path,
            title_line: tuple,
            field_keys: tuple,
            id_: bool,
            old=None,
            name="Solo_Download",
            *args,
            **kwargs):
        super().__init__(*args, **kwargs)
        self.file = None  # 文件对象
        self.writer = None  # CSV对象
        self.name = self._rename(root, self.__type, old, name)  # 文件名称
        self.path = root.joinpath(f"{self.name}.{self.__type}")  # 文件路径
        self.title_line = title_line  # 标题行
        self.field_keys = field_keys
        self.index = 1 if id_ else 0

    def __enter__(self):
        self.file = self.path.open(
            "a",
            encoding="UTF-8-SIG" if system() == "Windows" else "UTF-8",
            newline="")
        self.writer = writer(self.file)
        self.title()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.file.close()

    def title(self):
        if getsize(self.path) == 0:
            # 如果文件没有任何数据，则写入标题行
            self.save(self.title_line[self.index:])

    def save(self, data, *args, **kwargs):
        self.writer.writerow(data)


class XLSXLogger(NoneLogger):
    """XLSX格式"""
    __type = "xlsx"

    def __init__(
            self,
            root: Path,
            title_line: tuple,
            field_keys: tuple,
            id_: bool,
            old=None,
            name="Solo_Download",
            *args,
            **kwargs):
        super().__init__(*args, **kwargs)
        self.book = None  # XLSX数据簿
        self.sheet = None  # XLSX数据表
        self.name = self._rename(root, self.__type, old, name)  # 文件名称
        self.path = root.joinpath(f"{self.name}.{self.__type}")
        self.title_line = title_line  # 标题行
        self.field_keys = field_keys
        self.index = 1 if id_ else 0

    def __enter__(self):
        self.book = load_workbook(
            self.path) if self.path.exists() else Workbook()
        self.sheet = self.book.active
        self.title()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.book.save(self.path)
        self.book.close()

    def title(self):
        if not self.sheet["A1"].value:
            # 如果文件没有任何数据，则写入标题行
            for col, value in enumerate(self.title_line[self.index:], start=1):
                self.sheet.cell(row=1, column=col, value=value)

    def save(self, data, *args, **kwargs):
        self.sheet.append(data)


class SQLLogger(NoneLogger):
    """SQLite保存数据"""

    def __init__(
            self,
            root: Path,
            db_name: str,
            title_line: tuple,
            title_type: tuple,
            field_keys: tuple,
            id_: bool,
            old=None,
            name="Solo_Download", ):
        super().__init__()
        self.db = None  # 数据库
        self.cursor = None  # 游标对象
        self.name = (old, name)  # 数据表名称
        self.file = db_name  # 数据库文件名称
        self.path = root.joinpath(self.file)
        self.title_line = title_line  # 数据表列名
        self.title_type = title_type  # 数据表数据类型
        self.field_keys = field_keys
        self.index = 1 if id_ else 0

    def __enter__(self):
        self.db = connect(self.path)
        self.cursor = self.db.cursor()
        self.update_sheet()
        self.create()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.db.close()

    def create(self):
        create_sql = f"""CREATE TABLE IF NOT EXISTS {self.name} ({", ".join(
            [f"{i} {j}" for i, j in zip(self.title_line, self.title_type)])});"""
        self.cursor.execute(create_sql)
        self.db.commit()

    def save(self, data, *args, **kwargs):
        column = self.title_line[self.index:]
        insert_sql = f"""REPLACE INTO {self.name} ({", ".join(column)}) VALUES ({
        ", ".join(["?" for _ in column])});"""
        self.cursor.execute(insert_sql, data)
        self.db.commit()

    def update_sheet(self):
        old_sheet, new_sheet = self.name
        mark = new_sheet.split("_", 1)
        if not old_sheet or mark[-1] == old_sheet:
            self.name = new_sheet
            return
        mark[-1] = old_sheet
        old_sheet = "_".join(mark)
        update_sql = f"ALTER TABLE {old_sheet} RENAME TO {new_sheet};"
        self.cursor.execute(update_sql)
        self.db.commit()
        self.name = new_sheet


class RecordManager:
    """检查数据储存路径和文件夹"""
    works_keys = (
        "type",
        "collection_time",
        "uid",
        "sec_uid",
        "unique_id",
        "short_id",
        "id",
        "desc",
        "text_extra",
        "duration",
        "create_time",
        "nickname",
        "user_age",
        "signature",
        "downloads",
        "music_author",
        "music_title",
        "music_url",
        "origin_cover",
        "dynamic_cover",
        "tag_1",
        "tag_2",
        "tag_3",
        "digg_count",
        "comment_count",
        "collect_count",
        "share_count",
        "extra",
    )
    works_text = (
        "作品类型",
        "采集时间",
        "UID",
        "SEC_UID",
        "抖音号",
        "SHORT_ID",
        "作品ID",
        "作品描述",
        "作品话题",
        "视频时长",
        "发布时间",
        "账号昵称",
        "年龄",
        "账号签名",
        "作品地址",
        "音乐作者",
        "音乐标题",
        "音乐链接",
        "静态封面",
        "动态封面",
        "标签_1",
        "标签_2",
        "标签_3",
        "点赞数量",
        "评论数量",
        "收藏数量",
        "分享数量",
        "额外信息"
    )
    works_type = (
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT PRIMARY KEY",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "INTEGER",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "INTEGER",
        "INTEGER",
        "INTEGER",
        "INTEGER",
        "TEXT",
    )
    comment_keys = (
        "collection_time",
        "cid",
        "create_time",
        "uid",
        "sec_uid",
        "short_id",
        "unique_id",
        "nickname",
        "signature",
        "user_age",
        "ip_label",
        "text",
        "sticker",
        "image",
        "digg_count",
        "reply_comment_total",
        "reply_id",
        "reply_to_reply_id",
    )
    comment_title = (
        "采集时间",
        "评论ID",
        "评论时间",
        "UID",
        "SEC_UID",
        "SHORT_ID",
        "抖音号",
        "账号昵称",
        "账号签名",
        "年龄",
        "IP归属地",
        "评论内容",
        "评论表情",
        "评论图片",
        "点赞数量",
        "回复数量",
        "回复ID",
        "回复对象",
    )
    comment_type = (
        "TEXT",
        "TEXT PRIMARY KEY",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "INTEGER",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "INTEGER",
        "INTEGER",
        "TEXT",
        "TEXT",
    )
    search_user_keys = (
        "collection_time",
        "uid",
        "sec_uid",
        "nickname",
        "unique_id",
        "short_id",
        "avatar",
        "signature",
        "verify",
        "enterprise",
        "follower_count",
        "total_favorited",
    )
    search_user_title = (
        "采集时间",
        "UID",
        "SEC_UID",
        "账号昵称",
        "抖音号",
        "SHORT_ID",
        "头像链接",
        "账号签名",
        "标签",
        "企业",
        "粉丝数量",
        "获赞数量",
    )
    search_user_type = (
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "INTEGER",
        "INTEGER",
    )
    LoggerParams = {
        "works": {
            "db_name": "WorksData.db",
            "title_line": works_text,
            "title_type": works_type,
            "field_keys": works_keys,
            "id_": False,
        },
        "comment": {
            "db_name": "CommentData.db",
            "title_line": comment_title,
            "title_type": comment_type,
            "field_keys": comment_keys,
            "id_": False,
        },
        "search_user": {
            "db_name": "SearchData.db",
            "title_line": search_user_title,
            "title_type": search_user_type,
            "field_keys": search_user_keys,
            "id_": False,
        },
    }
    DataLogger = {
        "csv": CSVLogger,
        "xlsx": XLSXLogger,
        "sql": SQLLogger,
    }

    def run(
            self,
            parameter,
            folder="",
            type_="works",):
        root = parameter.root.joinpath(parameter.cleaner.filter_name(folder, False, "Data"))  # 文件存储路径
        root.mkdir(exist_ok=True)
        params = self.LoggerParams[type_]  # 对应参数体
        logger = self.DataLogger.get(parameter.storage_format, NoneLogger)  # 对应本地保存器
        return root, params, logger
